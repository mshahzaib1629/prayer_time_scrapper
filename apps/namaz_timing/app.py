import concurrent.futures
from openpyxl import Workbook
from apps.namaz_timing.utils.muslim_pro_prayer import scrap_prayer_timing_page
from apps.namaz_timing.utils.constants import month_names
import os
from openpyxl.styles import Font

class App:
    def __init__(self, city_name):
        
        # Create 'outputs' folder if it doesn't exist
        output_folder = "outputs"
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        self.excel_path = os.path.join(output_folder, f"{city_name}.xlsx")
        
        self.month_data = {}
        self.city_name = city_name
        
    def _write_to_excel(self):
        wb = Workbook()
        sheet = wb.active
        sheet.title = self.city_name
        
        odd_start_row = 1   # Starting row for odd months
        even_start_row = 38  # Starting row for even months
        columns_per_month = 9  # Number of columns each month's data takes
        
        for month_index, data in sorted(self.month_data.items()):
            month_name = month_names[month_index]
            is_odd_month = month_index % 2 == 1
            
            # Calculate starting positions
            month_position = (month_index + 1) // 2 - 1  # 0-based position for horizontal layout
            start_col = month_position * columns_per_month + 1  # Convert to 1-based column index
            current_row = odd_start_row if is_odd_month else even_start_row
            
            # Write month name as header
            bold_font = Font(bold=True)
            sheet.cell(row=current_row, column=start_col, value=month_name).font = bold_font
            
            # Write column headers and data
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 0):
                sheet.cell(row=current_row + 1, column=start_col + col, value=header)
            
            # Write data rows
            for row_offset, entry in enumerate(data, 2):
                for col, key in enumerate(headers, 0):
                    sheet.cell(row=current_row + row_offset, 
                             column=start_col + col, 
                             value=entry[key])
        
        wb.save(self.excel_path)
    
    def get_namaz_timings(self):
        with concurrent.futures.ThreadPoolExecutor() as executor:
            for month_index, tabledata in scrap_prayer_timing_page(self.city_name):
                month_name = month_names[month_index]
                
                self.month_data[month_index] = tabledata
                print(f"Page {month_name} processed!")
                print("------------------------------------------------------")
            
            print("Writing all data to Excel...")
            self._write_to_excel()
            print(f"Excel file saved at: {self.excel_path}")

